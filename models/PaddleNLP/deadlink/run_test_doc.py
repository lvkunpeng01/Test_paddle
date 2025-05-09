# -*- coding: utf-8 -*-
#!/usr/bin/env python
"""
some common func
"""

import sys
import os
import shutil
import difflib
import openpyxl
import requests
from bs4 import BeautifulSoup
import re
import time
import json
#import pymysql
import platform
import subprocess
import conf
import gc


# import test_sample
# import test_sample_api_en
def diff_line(file1, file2):
    """
    对两个文件做diff
    """
    diff_dict = {}
    with open(file1, "r", encoding="utf-8") as fr1:
        file1_lines = fr1.readlines()
    with open(file2, "r", encoding="utf-8") as fr2:
        file2_lines = fr2.readlines()
    file1_lines_set = set(file1_lines)
    file2_lines_set = set(file2_lines)
    file1_have_only = file1_lines_set - file2_lines_set
    file2_have_only = file2_lines_set - file1_lines_set
    diff_dict['file1_have_only'] = file1_have_only
    print('file1_have_only：%d' % len(file1_have_only))
    for file1 in file1_have_only:
        print(file1)
    diff_dict['file2_have_only'] = file2_have_only
    print('file2_have_only：%d' % len(file2_have_only))
    for file2 in file2_have_only:
        print(file2)
    # print(diff_dict)
    return diff_dict


def is_chinese(string):
    """
    检查整个字符串是否包含中文
    :param string: 需要检查的字符串
    :return: bool
    True表示中文
    """
    for ch in string:
        if u'\u4e00' <= ch <= u'\u9fff':
            return True
    return False


"""
def db_plus_valuelist(valuelist):
    ""
    为了更好的插入数据库，加引号，转义
    :param valuelist:
    :return:
    ""
    valuelist_plus = []
    for value in valuelist:
        if value is None:
            value = ''
        else:
            value = pymysql.escape_string(value)
        valuelist_plus.append("'%s'" % value)
    return valuelist_plus


def db_execute_sqllist(sql_list):
    ""
    执行sql列表
    :param sql_list:
    :return:
    ""
    select_flag = False
    conn = pymysql.connect(host=conf.MYSQL_HOSTNAME,
                           user=conf.MYSQL_USERNAME,
                           passwd=conf.MYSQL_PASSWORD,
                           db=conf.MYSQL_DATABASE,
                           port=conf.MYSQL_PORT,
                           charset='utf8')
    cursor = conn.cursor()
    rows_merge = ()
    for sql in sql_list:
        try:
            cursor.execute(sql)
            if sql.split(' ')[0] == 'select':
                rows = cursor.fetchall()
                rows_merge += rows
                select_flag = True
        except Exception as err:
            # 发生错误时回滚
            conn.rollback()
            exit(1)
    conn.commit()
    cursor.close()
    conn.close()
    if select_flag is True:
        return rows_merge
"""

def read_excel(excel_full, sheet):
    """
    根据文件全面、sheet名，读取Excel
    Args:
        excel_full_name:
        sheet:
    Returns:
    """
    wb = openpyxl.load_workbook(excel_full, read_only=True)
    ws = wb[sheet]
    rows = ws.rows
    lines = []
    for row in rows:
        line = [col.value for col in row]
        lines.append(line)
    del wb, ws
    gc.collect()
    return lines


def write_excel(path, value):
    """
    写excle
    :param path:
    :param value:
    :return:
    """
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = '工作表1'
    for i in range(0, len(value)):
        for j in range(0, len(value[i])):
            sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
    wb.save(path)
    del wb
    gc.collect()
    print("写入Excel成功！")


def get_file_list_filter(path, filter):
    """
    按照后缀名获取文件名(递归文件夹)
    :param path: 文件路径
    :param filter: 过滤文件（如"rst"）
    :return:
    """
    file_list = []
    lsdir = os.listdir(path)
    dirs = [i for i in lsdir if os.path.isdir(os.path.join(path, i))]
    files = [i for i in lsdir if os.path.isfile(os.path.join(path, i))]
    if dirs:
        for d in dirs:
            file = get_file_list_filter(os.path.join(path, d), filter)
            if file is not None:
                file_list.extend(file)
    if files:
        for f in files:
            if f.endswith(filter):
                file_list.append(os.path.join(path, f))
    return file_list


def read_file(file_name):
    """
    读取文件，分割，方便context diff
    Args:
        file_name:
    Returns:
    """
    try:
        file_desc = open(file_name, 'r')
        # 读取后按行分割
        text = file_desc.read().splitlines()
        file_desc.close()
        return text
    except IOError as error:
        print('Read input file Error: {0}'.format(error))
        sys.exit()


def diff_file_words(file1, file2):
    """
    去掉空格和换行，比较diff
    没有diff返回：''
    有diff返回：
    Args:
        file1:
        file2:
    Returns:
    """
    with open(file1, "r", encoding="utf-8") as fr1, \
            open(file2, "r", encoding="utf-8") as fr2:
        file1_words = fr1.read()
        file2_words = fr2.read()
        file1_words = file1_words.replace(" ", "").replace("\n", "")
        file2_words = file2_words.replace(" ", "").replace("\n", "")
    result = ""
    if file1_words != file2_words:
        diff = difflib.Differ().compare(file1_words.splitlines(), file2_words.splitlines())
        result = "\n".join(list(diff))
    return result


def diff_file_context(file1, file2):
    """
    比较两个文件内容，类似linux的diff展示
    能判断是否有diff
    Args:
        file1:
        file2:
    Returns:
    """
    # .splitlines(True)为保留了\r\n
    a = open(file1, "r").read().splitlines(True)
    # print(a)
    b = open(file2, "r").read().splitlines(True)
    # print(b)
    # difflib.context_diff(a, b[, fromfile][, tofile][, fromfiledate][, tofiledate][, n][, lineterm])
    # fromfile：a的文件名
    # tofile：b的文件名
    # fromfiledate：a文件的修改时间
    # tofiledate：b文件的修改时间
    # n：变化那一行前后n行展示，默认n=3
    # lineterm：打印出来的换行符，默认为\n
    diff = difflib.context_diff(a, b, fromfile=file1, tofile=file2, n=0, lineterm="\n")
    result = "".join(diff)
    # print(file1)
    # print(result)
    return result


def diff_file_html(file1, file2):
    """
    比较两个文件并把结果生成一份html文本
    不能判断是否有diff
    Args:
        file1:
        file2:
    Returns:
    """
    if file1 == "" or file2 == "":
        print("文件路径不能为空：第一个文件的路径：{0}, 第二个文件的路径：{1} .".format(file1, file2))
        sys.exit()
    else:
        print("正在比较文件{0} 和 {1}".format(file1, file2))
    text1_lines = read_file(file1)
    text2_lines = read_file(file2)
    diff = difflib.HtmlDiff()  # 创建HtmlDiff 对象
    result = diff.make_file(text1_lines, text2_lines)  # 通过make_file 方法输出 html 格式的对比结果
    # 将结果写入到result_comparation.html文件中
    try:
        file1 = file1.replace("_cn", "").replace("_en", "").replace(".py", ".html")
        diff_path = "%s%s/%s" % (conf.DIFF_HTML_PATH, file1.split("/")[-3], file1.split("/")[-2])
        diff_file = "%s/%s" % (diff_path, file1.split("/")[-1])
        if not os.path.exists(diff_path):
            os.makedirs(diff_path)
        with open(diff_file, "w") as result_file:
            result_file.write(result)
            print("0==}==========> Successfully Finished\n")
    except IOError as error:
        print('写入html文件错误：{0}'.format(error))


def diff_dir_filenames(dir_a, dir_b):
    """
    文件夹加下文件名做diff（按照文件名字对比）
    Args:
        dir_a:a文件夹
        dir_b:b文件夹
    Returns:
        only_a:a文件夹有，b文件夹没有的文件list
        only_b:a文件夹没有，b文件夹有的文件list
        both:a、b文件夹都有的文件list
        both_same:a、b文件夹都有，内容相同
        both_diff:a、b文件夹都有，内容不同
    """
    rt_dict = {}
    content_same = set()
    content_diff = set()
    a_flies = set()
    b_flies = set()
    if os.path.exists(dir_a):
        a_flies = set(os.listdir(dir_a))
    if os.path.exists(dir_b):
        b_flies = set(os.listdir(dir_b))
    both_flies = a_flies & b_flies
    for file in both_flies:
        a_file = dir_a + file
        b_file = dir_b + file
        diff_result = diff_file_words(a_file, b_file)
        if diff_result == '':
            content_same.add(file)
        else:
            content_diff.add(file)
            if not os.path.exists(conf.DIFF_HTML_PATH):
                os.makedirs(conf.DIFF_HTML_PATH)
            diff_file_html(a_file, b_file)
    rt_dict['only_a'] = a_flies - b_flies
    rt_dict['only_b'] = b_flies - a_flies
    rt_dict['both_same'] = content_same
    rt_dict['both_diff'] = content_diff
    rt_dict['dir_a'] = dir_a
    rt_dict['dir_b'] = dir_b
    return rt_dict


# ***【分界线以上为通用函数】******************************************************************
def get_doc_owner(excel_full, sheet):
    """
    获取doc的owner,通过"url key"判断
    Args:
        excel_full:
        sheet:
    Returns:
    """
    doc_owner = {}
    doc_datas = read_excel(excel_full, sheet)
    for doc_data in doc_datas:
        doc_url = doc_data[0]
        owner_cn = doc_data[1]
        if owner_cn is None:
            owner_cn = ''
        if doc_url is not None and doc_url.startswith('http'):
            key = doc_url.strip().split('zh/develop')[-1].split('.')[0].replace('_cn', '').replace('_en', '')
            api_dict = {}
            api_dict['key'] = key
            api_dict['owner'] = owner_cn
            doc_owner[key] = api_dict
    return doc_owner


def get_api_owner(excel_full, sheet):
    """
    获取api的owner
    Args:
        excel_full:
        sheet:
    Returns:
    """
    api_owner = {}
    api_datas = read_excel(excel_full, sheet)
    for api_data in api_datas:
        api_name = api_data[1]
        is_public = api_data[3]
        owner_cn = api_data[8]
        owner_email = api_data[9]
        if api_name is not None and is_public == 'public':
            api_name = api_name.strip()
            word_num = api_name.split('.')
            if len(word_num) == 3:
                key = "api%s" % api_name.replace('paddle', '')
            elif len(word_num) == 4:
                key = "api%s" % api_name.replace('paddle.fluid', '')
            elif len(word_num) == 2:
                key = "api.%s" % api_name
            api_dict = {}
            api_dict['key'] = key
            api_dict['api_name'] = api_name
            api_dict['owner'] = owner_cn
            if owner_email is not None:
                api_dict['owner'] = owner_email
            api_owner[key] = api_dict
    return api_owner


def get_api_owner_new(excel_full, sheet):
    """
    获取api的owner
    Args:
        excel_full:
        sheet:
    Returns:
    """
    api_owner = {}
    api_datas = read_excel(excel_full, sheet)
    for api_data in api_datas:
        api_name = api_data[1]
        is_public = api_data[3]
        owner_cn = api_data[8]
        owner_email = api_data[9]
        if api_name is not None and is_public == 'public':
            api_name = api_name.strip()
            if len(api_name.split('.')) == 3:
                key = "api%s" % api_name.replace('paddle', '')
            else:
                key = "api%s" % api_name.replace('paddle.fluid', '')
            api_dict = {}
            key_new = "/%s" % key.replace('.', '/')
            api_dict['key'] = key_new
            api_dict['api_name'] = api_name
            api_dict['owner'] = owner_cn
            if owner_email is not None:
                api_dict['owner'] = owner_email
            api_owner[key_new] = api_dict
    return api_owner


def file_to_link(file, version):
    """
    将路径转化为html
    :param file: 文件名带路径
    :param version: 语言（en:英语，zh:中文）/版本（默认""表示当前版本，1.7、1.6、...）
    :return:
    """
    link_dict = {}
    file_end = file.split("/")[-1].replace(".rst", "")
    if file_end.endswith("_cn"):
        lan_version = "zh/%s" % version
    else:
        lan_version = "en/%s" % version
    path_end = file.split("doc/fluid/")[-1].replace(".rst", "")
    link = "%s/documentation/docs" \
           "/%s/%s%s" % (conf.PADDLE_HOST, lan_version, path_end, ".html")
    key = path_end.replace("_cn", "").replace("_en", "").replace("/", ".")
    link_dict['key'] = key
    link_dict['lan_version'] = lan_version
    link_dict['link'] = link
    link_dict['file'] = file
    link_dict['path'] = path_end.rsplit('/', 1)[0]
    return link_dict


def key_to_link(key, version):
    """
    通过key推算中英文链接
    Args:
        key:
        version:
    Returns:
    api.layers.create_py_reader_by_data
    """
    result = {}
    key_list = key.split('.')
    lan_version_cn = 'zh/%s' % version
    lan_version_en = 'en/%s' % version
    if key_list[0] == 'api':
        path_end_cn = '%s_cn/%s_cn/%s_cn' % (key_list[0], key_list[1], key_list[2])
        path_end_en = '%s/%s/%s' % (key_list[0], key_list[1], key_list[2])
    link_cn = '%s/documentation/docs' \
              '/%s/%s%s' % (conf.PADDLE_HOST, lan_version_cn, path_end_cn, '.html')
    link_en = '%s/documentation/docs' \
              '/%s/%s%s' % (conf.PADDLE_HOST, lan_version_en, path_end_en, '.html')
    result['link_cn'] = link_cn
    result['link_en'] = link_en
    return result


def samplefile_to_name(file):
    """
    给出路径，获取key
    Args:
        file:fluid_sample/api/framework/Variable.persistable_1.py
    Returns:
        Variable
    """
    file_1 = file.split('/')[-1]
    file_2 = file_1.rsplit('_', 1)[0]
    file_3 = file_2.split('.')[0]
    return file_3


def diff_dir_context(diff_result):
    """
    进行二次比对，按照内容一一对比
    Args:
        diff_result:
    Returns:
        only_a:a有，b没有的文件set
        only_b:a没有，b有的文件set
        both:a、b文件夹都有的文件set
        both_same:a、b都有，内容相同
        both_diff:a、b都有，内容不同
    """
    rt_dict = {}
    content_same_a = set()
    content_same_b = set()
    fileset_a = diff_result['only_a'] | diff_result['both_diff']
    fileset_b = diff_result['only_b'] | diff_result['both_diff']
    for b_file in fileset_b:
        for a_file in fileset_a:
            a_key = samplefile_to_name(a_file)
            b_key = samplefile_to_name(b_file)
            if a_key == b_key:
                file_words = diff_file_words(diff_result['dir_a'] + a_file,
                                             diff_result['dir_b'] + b_file)
                # 如果有相同的，跳出循环
                if file_words == '':
                    content_same_a.add(a_file)
                    content_same_b.add(b_file)
                    break
    rt_dict['only_a'] = fileset_a - content_same_a - diff_result['both_diff']
    rt_dict['only_b'] = fileset_b - content_same_b - diff_result['both_diff']
    rt_dict['both_same'] = {'content_same_a': (diff_result['both_same'] | content_same_a),
                            'content_same_b': (diff_result['both_same'] | content_same_b)}
    rt_dict['both_diff'] = diff_result['both_diff'] - content_same_a
    rt_dict['dir_a'] = diff_result['dir_a']
    rt_dict['dir_b'] = diff_result['dir_b']
    return rt_dict


def write_icafe_batch(excel_table):
    """
    写入icafe，分批写，每20个case写一次
    Args:
        excel_table:
        # [['相对路径', '代码段', 'diff状态', 'owner', 'key',
        '中文链接', '英文链接', '版本', 'diff链接']]
    Returns:
    """
    # 分批创建issue，避免icafe api超时
    icafe_dict_list = []
    i = 0
    # 将buglist从小到大排序
    for row in excel_table:
        if i % 20 == 0:
            icafe_dict = {'username': conf.ICAFE_USERNAME, 'password': conf.ICAFE_PASSWORD}
            issue_list = []
        item_dict = {}
        result = row[2]
        owner = row[3]
        # item_dict['detail'] = "辛苦填写：1.用户场景简述；2.问题根因；3.改进措施。"
        item_dict['type'] = 'Bug'
        item_dict['title'] = '【%s版本】【%s】中代码段%s存在%s问题' % (row[7], row[4], row[1], result)
        item_dict["fields"] = {
            'owner': owner,
            '中文链接': row[5],
            '英文链接': row[6]
        }
        if result == 'sample code diff':
            item_dict['fields']['diff链接'] = row[8]
        # if owner != '' and owner != 'None':
        #     item_dict['fields']['负责人'] = owner.strip()
        issue_list.append(item_dict)
        if (i % 20 == 19) or (i == len(excel_table) - 1):
            icafe_dict['issues'] = issue_list
            icafe_dict_list.append(icafe_dict)
        i += 1
    for i, icafe_dict in enumerate(icafe_dict_list):
        result = requests.post(conf.ICAFE_API_NEWCARD, data=json.dumps(icafe_dict))
        result = json.loads(result.text)
        print('写入Icafe【%s】：%s' % (result['status'], result['message']))
        time.sleep(1)
    return icafe_dict_list


def write_icafe(excel_table):
    """
    写入icafe，一个一个写，方便看哪个负责人有问题
    Args:
        excel_table:
        # [['相对路径', '代码段', 'diff状态', 'owner', 'key',
        '中文链接', '英文链接', '版本', 'diff链接']]
    Returns:
    """
    # 将buglist从小到大排序
    for row in excel_table:
        icafe_dict = {'username': conf.ICAFE_USERNAME, 'password': conf.ICAFE_PASSWORD}
        item_dict = {}
        diff_type = row[2]
        owner = row[3]
        # item_dict['detail'] = "辛苦填写：1.用户场景简述；2.问题根因；3.改进措施。"
        item_dict['type'] = 'Bug'
        item_dict['title'] = '【%s版本】【%s】中代码段%s存在%s问题' % (row[7], row[4], row[1], diff_type)
        item_dict["fields"] = {
            'owner': owner,
            '中文链接': row[5],
            '英文链接': row[6],
            '需求来源': 'QA团队'
        }
        # diff_type
        if diff_type == 'sample code diff':
            item_dict['fields']['diff链接'] = row[8]
        # 负责人有异常的时候
        item_dict['fields']['负责人'] = owner.strip()
        # 一个一个写，所以只装一个item_dict
        icafe_dict['issues'] = [item_dict]
        result = requests.post(conf.ICAFE_API_NEWCARD, data=json.dumps(icafe_dict))
        result_dict = json.loads(result.text)
        if result_dict['status'] == 401:
            item_dict['fields'].pop('负责人')
            icafe_dict['issues'] = [item_dict]
            print('owner error，查无此人（或有重名）：%s' % owner)
            result = requests.post(conf.ICAFE_API_NEWCARD, data=json.dumps(icafe_dict))
            result_dict = json.loads(result.text)
        print('写入Icafe【%s】：%s' % (result_dict['status'], result_dict['message']))
        time.sleep(1)


def test_diff_lan_api(version):
    """
    对比中英文api sample diff
    1.中文有、英文没有的文档
    2.中文没有、英文有的文档
    3.中英文文档都有、文档内样例数目相等，但是有样例diff
    5.完全一致
    Returns:
    excel: 相对路径，代码段，diff状态，负责人，diff对比地址，页面a，页面b
    """
    # 清理环境
    if os.path.exists(conf.DIFF_HTML_PATH):
        shutil.rmtree(conf.DIFF_HTML_PATH)
    api_cn_dir_set = set()
    api_en_dir_set = set()
    excel_table_new = [['相对路径', '代码段', 'diff状态', 'owner', 'key',
                        '中文链接', '英文链接', '版本', 'diff链接']]
    excel_table = []
    icafe_list = []
    # 查看所有文件夹和子文件夹
    api_cn = conf.RESULT_SAMPLE_PATH + 'api_cn/'
    api_en = conf.RESULT_SAMPLE_PATH + 'api/'
    for root, dirs, files in os.walk(api_cn):
        for dir in dirs:
            api_cn_dir_set.add(dir.replace('_cn', ''))
    for root, dirs, files in os.walk(api_en):
        for dir in dirs:
            api_en_dir_set.add(dir)
    # 并集，方便按照文件夹名、文件名逐一diff
    dirs_set = api_cn_dir_set | api_en_dir_set
    for dir in dirs_set:
        dir_a = '%s%s%s/' % (api_cn, dir, '_cn')
        dir_b = '%s%s/' % (api_en, dir)
        # dir_a = 'fluid_sample/api_cn/executor_cn/'
        # dir_b = 'fluid_sample/api/executor/'
        diff_dir_result = diff_dir_filenames(dir_a, dir_b)
        diff_result = diff_dir_context(diff_dir_result)
        path = dir_a.split(conf.RESULT_SAMPLE_PATH)[-1].replace('_en', '').replace('_cn', '')
        for file in diff_result['only_a']:
            excel_line = [path, file, 'sample only exist in Chinese']
            excel_table.append(excel_line)
        for file in diff_result['only_b']:
            excel_line = [path, file, 'sample only exist in English']
            excel_table.append(excel_line)
        for file in diff_result['both_diff']:
            excel_line = [path, file, 'sample code diff']
            excel_table.append(excel_line)
        for file in diff_result['both_same']['content_same_a']:
            excel_line = [path, file, 'sample same']
            excel_table.append(excel_line)
    # 给每一行加owner
    api_owner = get_api_owner(conf.OWNER_API_PATH, conf.OWNER_API_SHEET)
    for line in excel_table:
        path = line[0]
        file = line[1]
        diff_state = line[2]
        diff_html = ''
        if diff_state == 'sample code diff':
            diff_html = ('%s%s/%s%s' % (conf.DIFF_HTML_HOST, version, path, file.replace('.py', '.html')))
        key_sub = path.replace('/', '.')
        key_end = file[:file.rfind('_')].split('.')[0]
        key = '%s%s' % (key_sub, key_end)
        owner = ''
        if key in api_owner.keys():
            owner = api_owner[key]['owner']
        line.append(owner)
        line.append(key)
        link = key_to_link(key, version)
        line.append(link['link_cn'])
        line.append(link['link_en'])
        line.append(version)
        line.append(diff_html)
        excel_table_new.append(line)
        if diff_state != 'sample same':
            icafe_list.append(line)
    write_excel(conf.RESULT_EXCEL_FILE, excel_table_new)
    # write_icafe(icafe_list)
    return icafe_list


# 匹配 python 的div代码,同一个页面中的代码写进一个py,用于教程类示例的验证
def spider_doc_sample_tutorial(url, sample_path, python_div):
    """
    抓取html存起来，去除注释和空格，并写进py中，但是不同div会写进不同文件，待改进
    Args:
        url: 官网链接
        sample_path: 存样例的地址
    Returns:
    """
    # 抓取
    try:
        # 设置TCP连接超时为3s，HTTP请求超时为7s
        r = requests.get(url, timeout=(3, 20))
        # print("spider_doc_sample-----try", r.text)
    except Exception as err:
        # 发生错误时回滚
        print('url请求失败：%s', url)
        # exit(1)
    if not os.path.exists(sample_path):
        os.makedirs(sample_path)
    # 样例带路径前缀（后面会根据样例位置添加1、2...）
    doc_name = url.split('/')[-1].replace('.html', '')
    if doc_name.endswith("_cn"):
        doc_name = doc_name.replace("_cn", "")
    # 分析，写文件
    soup = BeautifulSoup(r.text, 'html.parser')
    # print("------soup----\n", soup)
    # samples = soup.find_all(class_=re.compile(r'highlight highlight-source-python'))
    samples = soup.find_all(class_=re.compile(python_div))
    print("---spider_doc_sample_tutorial---url----", url)
    # 由于slim md教程的特性是一个md文档中的代码是一个示例,因此需要写在同一个脚本中;
    sample_name = '%s/%s.py' % (sample_path, doc_name)
    print("--sample_name--", sample_name)
    for i, sample in enumerate(samples):
        sample_lines = sample.get_text().split('\n')
        # print("--sample_lines--", sample_lines)
        code_first_line = sample_lines[0]
        # 当代码段里有中文等不存储,还需要去掉标点符号的行
        if (is_chinese(code_first_line) is True) or \
                ('Given:' in code_first_line) or ('Case' in code_first_line) \
                or ('Example:' in code_first_line) or ('!rm' in code_first_line) \
                or ('cd:' in code_first_line) or ('python:' in code_first_line):
            continue

        with open(sample_name, 'a+') as fw:
            # 首行加import fluid
            # fw.write('import paddle.fluid as fluid\n')
            for line in sample_lines:
                line_strip = line.strip()
                if line_strip == '' or line_strip.startswith('#'):
                    continue
                else:
                    fw.write('%s\n' % line.split('#')[0])


# 匹配 python 的div代码,同一个页面中的代码写进不同py,用于验证API的示例验证
def spider_doc_sample_api(url, sample_path, python_div):
    """
    抓取html存起来，去除注释和空格，并写进py中，但是不同div会写进不同文件，待改进
    Args:
        url: 官网链接
        sample_path: 存样例的地址
    Returns:
    """
    # 抓取
    try:
        # 设置TCP连接超时为3s，HTTP请求超时为7s
        r = requests.get(url, timeout=(3, 7))
        # print("spider_doc_sample-----try", r.text)
    except Exception as err:
        # 发生错误时回滚
        print('url请求失败：%s', url)
        # exit(1)
    if not os.path.exists(sample_path):
        os.makedirs(sample_path)
    # 样例带路径前缀（后面会根据样例位置添加1、2...）
    doc_name = url.split('/')[-1].replace('.html', '')
    if doc_name.endswith("_cn"):
        doc_name = doc_name.replace("_cn", "")
    # 分析，写文件
    soup = BeautifulSoup(r.text, 'html.parser')
    # print("------soup----\n", soup)
    # samples = soup.find_all(class_=re.compile(r'highlight highlight-source-python'))
    samples = soup.find_all(class_=re.compile(python_div))
    # print("------samples----", samples)
    # 由于slim md教程的特性是一个md文档中的代码是一个示例,因此需要写在同一个脚本中;
    # 但是API 的示例还是要分开写
    for i, sample in enumerate(samples):
        sample_lines = sample.get_text().split('\n')
        code_first_line = sample_lines[0]
        # 当代码段里有中文等不存储
        if (is_chinese(code_first_line) is True) or \
                ('Given:' in code_first_line) or ('Case' in code_first_line) \
                or ('Example:' in code_first_line) or ('!rm' in code_first_line):
            continue
        sample_name = '%s/%s_%d.py' % (sample_path, doc_name, i + 1)
        with open(sample_name, 'w') as fw:
            # 首行加import fluid
            # fw.write('import paddle.fluid as fluid\n')
            for line in sample_lines:
                line_strip = line.strip()
                if line_strip == '' or line_strip.startswith('#'):
                    continue
                else:
                    fw.write('%s\n' % line.split('#')[0])


# 匹配 python 的div代码
def spider_doc_sample_bk(url, sample_path):
    """
    抓取html存起来，去除注释和空格
    Args:
        url: 官网链接
        sample_path: 存样例的地址
    Returns:
    """
    # 抓取
    try:
        # 设置TCP连接超时为3s，HTTP请求超时为7s
        r = requests.get(url, timeout=(3, 7))
        print("spider_doc_sample-----try", r.text)
    except Exception as err:
        # 发生错误时回滚
        print('url请求失败：%s', url)
        exit(1)
    if not os.path.exists(sample_path):
        os.makedirs(sample_path)
    # 样例带路径前缀（后面会根据样例位置添加1、2...）
    doc_name = url.split('/')[-1].replace('.html', '')
    if doc_name.endswith("_cn"):
        doc_name = doc_name.replace("_cn", "")
    # 分析，写文件
    soup = BeautifulSoup(r.text, 'html.parser')
    print("------soup----\n", soup)
    samples = soup.find_all(class_=re.compile(r'highlight-python.'))
    print("------samples----", samples)
    for i, sample in enumerate(samples):
        sample_lines = sample.get_text().split('\n')
        code_first_line = sample_lines[0]
        # 当代码段里有中文等不存储
        if (is_chinese(code_first_line) is True) or \
                ('Given:' in code_first_line) or ('Case' in code_first_line) \
                or ('Example:' in code_first_line):
            continue
        sample_name = '%s/%s_%d.py' % (sample_path, doc_name, i + 1)
        with open(sample_name, 'w') as fw:
            # 首行加import fluid
            fw.write('import paddle.fluid as fluid\n')
            for line in sample_lines:
                line_strip = line.strip()
                if line_strip == '' or line_strip.startswith('#'):
                    continue
                else:
                    fw.write('%s\n' % line.split('#')[0])


def run_sample_code(filename):
    """
    运行样例
    three status ,-1:no sample code; 1: running error; 0:normal
    :param filename:
    :return:
    """
    python_verion = platform.python_version()[0]
    python_path = '%s%s' % ('python', python_verion)
    # cmd = ['python3', filename]
    # cmd = ['python2', filename]
    cmd = [python_path, filename]
    subprc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    _, error = subprc.communicate()
    err_detail = ''.join(error.decode(encoding='utf-8'))
    if subprc.returncode != 0:
        print('\nSample code error found in ', filename, ':\n')
        print(err_detail)
        err_short = err_detail.split('\n')[-2]
        if ':' in err_short:
            err_short = err_short.split(':')[0]
        status = ['FAIL', err_short, err_detail]
    else:
        status = ['SUCCESS', '', '']
    return status


def run_samples_by_samplelist(file_list, result_file):
    """
    运行文件夹下所有的样例（.py）
    Args:
        file_list: 即样例地址
        result_file:
    Returns:
    """
    result_list = [
        ['NOTE：代码名 Program_5.py 表示 Program页面的第5个代码样例']
        , ['序号', '代码名', '测试结果', '错误类型', '错误详情', '负责人', '代码内容']
    ]
    result_txt = '%s.txt' % result_file
    result_excel = '%s.xlsx' % result_file
    # api_owner = get_api_owner(conf.OWNER_API_PATH, conf.OWNER_API_SHEET)
    # 存txt
    result_f = open(result_txt, 'a+')
    for i, file in enumerate(file_list):
        print('[%d-%d] python %s' % (len(file_list), i, file))
        result = run_sample_code(file)
        item_list = []
        item_list.append(str(i))
        item_list.append(file)
        item_list.extend(result)
        if result[0] == 'FAIL':
            # 如果有错误，附上代码
            with open(file, 'r') as fr:
                sample_code = fr.read()
            # 如果第一行含有中文或不规范的注释，则不记录
            code_first_line = sample_code.split('\n')[1]
            if (is_chinese(code_first_line) is True) or \
                    ('Given:' in code_first_line) or ('Case' in code_first_line) \
                    or ('Example:' in code_first_line):
                continue
            # 增加负责人
            tmp_1 = file[:file.rfind('_')]
            tmp_2 = tmp_1.split(sample_path)[-1]
            tmp_3 = sample_path.split('/')[-2]
            key = ('%s/%s' % (tmp_3, tmp_2)).replace('/', '.').replace('_cn', '').replace('_en', '')
            # owner = ''
            # if key in api_owner.keys():
            #     owner = api_owner[key]['owner']
            # if owner is None:
            #     owner = ''
            # item_list.append(owner)
            # 增加样例代码
            item_list.append(sample_code)
            result_list.append(item_list)
        result_f.write("\t".join(item_list) + "\n")
    result_f.close()
    write_excel(result_excel, result_list)


def run_samples(sample_path, result_file):
    """
    运行文件夹下所有的样例（.py）
    Args:
        sample_path:
        result_file:
    Returns:
    """
    file_list = get_file_list_filter(sample_path, ".py")
    print(file_list)
    run_samples_by_samplelist(file_list, result_file)


# 暂不需要
def run_samples_again_by_excel(result_file_old, result_file_new):
    """
    对python3下跑不通的在python2下继续跑，如果跑通了则更新结果
    Args:
        sample_path:
        result_file:
    Returns:
    """
    file_list = []
    results_old = read_excel(result_file_old, '工作表1')
    for i, result_old in enumerate(results_old):
        if i == 0:
            continue
        file_list.append(result_old[2])
    print(file_list)
    run_samples_by_samplelist(file_list, result_file_new)


def load_white():
    """
    load white list
    """
    LIST_WHITE = []
    with open("LIST.WHITE", "r") as fin:
        lines = fin.readlines()
        for line in lines:
            LIST_WHITE.append(line.strip())
    return LIST_WHITE


if __name__ == "__main__":
    sample_path = 'paddle_sample/api_cn'
    run_samples(sample_path, conf.RESULT_SAMPLE_PATH + 'result_run_slim_api_cn')
